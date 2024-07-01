import requests,bs4,re,time
from urllib.parse import urlparse ,urlunparse,parse_qs,urlencode
from openpyxl import Workbook

#  start of variable assosiated with the  write to excel, write to excel should be written on its own file later
x = []
wb = Workbook()
sheet = wb.active

pc= 1
prc=2
lc=3

def write_to_excel(data):
    row = 1
    for item in data:
        sp = item.split(',')
        cur = {
            "product":sp[0],
            "price": "XXX",
            "link":"XXX"
        }
        x.append(cur)
        sheet.cell(row,pc).value= cur['product']
        sheet.cell(row,prc).value= cur['price']
        sheet.cell(row,lc).value = cur['link']
        row += 1
    wb.save('Laptops.xlsx')


# end of the variables

filtered_data=[]
url=''
ram_regex = re.compile(r"\b(\d+)\s?GB\s?RAM\b")

def edit_url_query(n):
    u_parse = urlparse('https://www.jumia.co.ke/laptops/#catalog-listing')
    q_dictionary = parse_qs(u_parse.query)
    q_dictionary["page"] = n
    rejoined_query = urlencode(q_dictionary) 
    rejoined_url =urlunparse(u_parse._replace(query = rejoined_query))
    return rejoined_url




iterator = 1
while iterator <= 50:
    if iterator == 1:
        url = 'https://www.jumia.co.ke/laptops/#catalog-listing'
    else:
        url = edit_url_query(iterator)
    
    req = requests.get(url)
    req.raise_for_status()
    print('Scrapping page ........................................................')

    passed_html = bs4.BeautifulSoup(req.text,'html.parser')
    scrapped_data = passed_html.select('h3.name')

    for product in scrapped_data:
        mo = ram_regex.search(str(product))
        if mo:
            ram_value = int(mo.group(1))
            if ram_value >=8:
                
                filtered_data.append(product.getText())
    iterator +=1

write_to_excel(filtered_data)




        


